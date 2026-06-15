# Generic export utilities for Excel and PDF reports across modules.
# Used for: Commission exports (SAL-08), Financial reports (FIN-09), etc.

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse
from datetime import datetime


def export_to_excel(data: list, columns: list, title: str = "Report") -> StreamingResponse:
    """
    data: list of dicts
    columns: list of (key, header_label) tuples defining column order and headers
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_fill = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(data, start=2):
        for col_idx, (key, label) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{title.replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def export_to_pdf(data: list, columns: list, title: str = "Report") -> StreamingResponse:
    """
    data: list of dicts
    columns: list of (key, header_label) tuples defining column order and headers
    Generates a simple tabular PDF report using reportlab.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    headers = [label for _, label in columns]
    table_data = [headers]
    for item in data:
        row = [str(item.get(key, "")) for key, _ in columns]
        table_data.append(row)

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B3A6B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FA")]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    filename = f"{title.replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )